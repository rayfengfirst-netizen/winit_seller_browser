"""
定时任务入口（建议 05:00）：按账号执行 inventoryFlow 导出，下载 InventoryInoutSeller 文件，
解压并读取表格后入库（覆盖写入，不保留每日历史），并推送飞书摘要。

重要：
- 使用独立数据库（WINIT_INOUT_SQLITE_PATH / 默认 artifacts/winit_inout.db）
- 不与老库存链路 run_daily_winit_job.py 共用表
"""

from __future__ import annotations

import csv
import json
import os
import subprocess
import sys
import tempfile
import zipfile
from datetime import date, datetime, timezone
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
load_dotenv(ROOT / ".env")

from winit_accounts import (  # noqa: E402
    WinitAccount,
    downloads_dir_base,
    list_winit_accounts,
    resolve_download_dir_for_account,
)
from winit_feishu_webhook import feishu_send_text  # noqa: E402
from winit_inventory_inout_db import (  # noqa: E402
    connect_inout,
    init_inout_schema,
    inout_sqlite_path,
    replace_inout_current_rows,
    upsert_inout_latest_meta,
)


ENTRY_URL = "https://seller.winit.com.cn/Australia/inventoryFlow"
ROW_MATCH = "InventoryInoutSeller"


def _inout_report_url() -> str:
    base = os.environ.get("WINIT_PUBLIC_BASE_URL", "").strip().rstrip("/")
    if not base:
        return ""
    return f"{base}/report/inout-shelf"


def _snapshot_date_str() -> str:
    raw = os.environ.get("WINIT_SNAPSHOT_DATE", "").strip()
    if raw:
        return raw
    return date.today().isoformat()


def _run_step02_inout_for_account(account: WinitAccount) -> int:
    """
    以子进程运行 step02，避免在当前进程污染环境变量，确保与原库存链路互不干扰。
    """
    env = os.environ.copy()
    env["WINIT_STEP02_ENTRY_URL"] = ENTRY_URL
    env["WINIT_EXPORT_ROW_MATCH"] = ROW_MATCH
    env["WINIT_STEP02_SKIP_SUCCESS_GATE"] = env.get("WINIT_STEP02_SKIP_SUCCESS_GATE", "1")
    env["WINIT_EXPORT_DIALOG_SKIP"] = env.get("WINIT_EXPORT_DIALOG_SKIP", "1")
    # inventoryFlow 常无库存页那种 radio，留空表示直接点确定。
    env["WINIT_EXPORT_DIALOG_RADIO_LABEL"] = env.get("WINIT_EXPORT_DIALOG_RADIO_LABEL", "")
    env["WINIT_ACCOUNT_ID"] = str(account.id)
    env["WINIT_RUN_ALL_ACCOUNTS"] = "0"
    cp = subprocess.run(
        [sys.executable, str(ROOT / "step02_australia_export.py")],
        cwd=str(ROOT),
        env=env,
        check=False,
    )
    return int(cp.returncode)


def _iter_candidate_files(d: Path):
    pats = [
        "*InventoryInoutSeller*.zip",
        "*InventoryInoutSeller*.xlsx",
        "*InventoryInoutSeller*.xls",
        "*InventoryInoutSeller*.csv",
    ]
    for pat in pats:
        for p in d.glob(pat):
            yield p


def find_latest_inout_file_for_account(account: WinitAccount) -> Path:
    d = resolve_download_dir_for_account(account)
    files = list(_iter_candidate_files(d))
    if not files and account.id == 1:
        root = downloads_dir_base()
        files = [p for p in _iter_candidate_files(root) if p.parent == root]
    if not files:
        raise FileNotFoundError(f"账号 {account.id} 目录下无 {ROW_MATCH} 文件：{d}")
    return max(files, key=lambda p: p.stat().st_mtime)


def _norm_header(h: object, idx: int) -> str:
    s = str(h).strip() if h is not None else ""
    return s if s else f"col_{idx}"


def _load_rows_from_xlsx(xlsx: Path) -> tuple[str, list[dict]]:
    wb = load_workbook(filename=str(xlsx), read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        sheet_name = ws.title or "Sheet1"
        rows = ws.iter_rows(values_only=True)
        try:
            head = next(rows)
        except StopIteration:
            return sheet_name, []
        headers = [_norm_header(h, i + 1) for i, h in enumerate(head)]
        out: list[dict] = []
        for row in rows:
            if not any(x not in (None, "") for x in row):
                continue
            d = {}
            for i, v in enumerate(row):
                k = headers[i] if i < len(headers) else f"col_{i+1}"
                d[k] = v
            out.append(d)
        return sheet_name, out
    finally:
        wb.close()


def _load_rows_from_csv(csv_path: Path) -> tuple[str, list[dict]]:
    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        return "csv", []
    headers = [_norm_header(h, i + 1) for i, h in enumerate(rows[0])]
    out: list[dict] = []
    for row in rows[1:]:
        if not any(x not in ("", None) for x in row):
            continue
        d = {}
        for i, v in enumerate(row):
            k = headers[i] if i < len(headers) else f"col_{i+1}"
            d[k] = v
        out.append(d)
    return "csv", out


def _load_rows_from_file(file_path: Path) -> tuple[str, list[dict]]:
    suf = file_path.suffix.lower()
    if suf in (".xlsx", ".xlsm"):
        return _load_rows_from_xlsx(file_path)
    if suf == ".csv":
        return _load_rows_from_csv(file_path)
    if suf == ".zip":
        with tempfile.TemporaryDirectory() as td:
            tdir = Path(td)
            with zipfile.ZipFile(file_path, "r") as zf:
                zf.extractall(tdir)
            xlsx = sorted(tdir.rglob("*.xlsx"))
            if xlsx:
                return _load_rows_from_xlsx(xlsx[0])
            csvs = sorted(tdir.rglob("*.csv"))
            if csvs:
                return _load_rows_from_csv(csvs[0])
    raise RuntimeError(f"不支持的文件类型或 zip 内无 xlsx/csv: {file_path}")


def run_one_account(conn, account: WinitAccount) -> tuple[int, int, str, str]:
    try:
        code = _run_step02_inout_for_account(account)
        if code != 0:
            return code, 0, "", f"step02 退出码 {code}"

        f = find_latest_inout_file_for_account(account)
        file_path_str = str(f)
        sheet_name, rows = _load_rows_from_file(f)
        rows_json = [json.dumps(r, ensure_ascii=False, default=str) for r in rows]
        row_count = replace_inout_current_rows(
            conn,
            account_id=account.id,
            account_username=account.username,
            file_name=f.name,
            sheet_name=sheet_name,
            rows_json=rows_json,
        )
        finished = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
        upsert_inout_latest_meta(
            conn,
            account_id=account.id,
            account_username=account.username,
            file_path=file_path_str,
            file_name=f.name,
            file_size_bytes=f.stat().st_size,
            row_count=row_count,
            updated_at=finished,
        )
        return 0, row_count, f.name, ""
    except Exception as e:
        return 1, 0, "", f"{type(e).__name__}: {e}"


def main() -> int:
    accs = list_winit_accounts()
    if not accs:
        print("未配置任何 Winit 账号（.env）", file=sys.stderr)
        feishu_send_text(
            "❌ InventoryInoutSeller 同步失败：未配置任何账号（.env）",
            channel="sync",
        )
        return 1

    snapshot_date = _snapshot_date_str()
    db_path = inout_sqlite_path()
    print(
        f"[inout] 日期 {snapshot_date}，账号数 {len(accs)}，入口 {ENTRY_URL}，匹配 {ROW_MATCH}",
        flush=True,
    )

    conn = connect_inout()
    init_inout_schema(conn)
    exit_code = 0
    lines: list[str] = [
        "📊 InventoryInoutSeller 下载入库",
        f"快照日期：{snapshot_date}",
        f"SQLite：{db_path}",
        "入库策略：按账号覆盖写入（不保留每日历史）",
        "",
    ]
    try:
        for account in accs:
            code, rows, fname, err = run_one_account(conn, account)
            if code != 0:
                exit_code = code
                lines.append(f"❌ 账号 {account.id}（{account.username}）失败：{err}")
            else:
                lines.append(f"✅ 账号 {account.id}（{account.username}）文件 {fname}，入库 {rows} 行")
    finally:
        conn.close()

    lines[0] = (
        "📊 InventoryInoutSeller 下载入库 ✅ 全部成功"
        if exit_code == 0
        else "📊 InventoryInoutSeller 下载入库 ⚠️ 有账号失败"
    )
    summary = "\n".join(lines)
    ok, detail = feishu_send_text(summary, channel="sync")
    print(f"[inout] 飞书(sync)：{'成功' if ok else '失败'} ({detail})", flush=True)

    if exit_code == 0:
        url = _inout_report_url()
        quick = "✅ 账号上架数据已更新，请速速查看。"
        if url:
            quick += f"\n链接：{url}"
        ok2, detail2 = feishu_send_text(quick, channel="inout_shelf")
        print(
            f"[inout] 飞书(inout_shelf)：{'成功' if ok2 else '失败'} ({detail2})",
            flush=True,
        )

    return exit_code


if __name__ == "__main__":
    raise SystemExit(main())
