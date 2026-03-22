"""
第三步：解压万邑通库存导出 zip（内为「海外仓库存_*_仓库级别_*.xlsx」），预览表头与样例行，可选导出 CSV。

默认在 downloads/ 下递归查找最新的 inventorySellerPortalExport-*.zip；也可显式传入 zip 路径。

运行：
  cd winit_seller_browser && source .venv/bin/activate
  python step03_unpack_winit_export.py
  python step03_unpack_winit_export.py downloads/LX/某文件.zip --export-csv
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import sys
import warnings
import zipfile
from pathlib import Path
from typing import Iterable, List, Optional

import openpyxl

# 万邑通导出 xlsx 常无默认样式，避免刷屏
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

ROOT = Path(__file__).resolve().parent


def _downloads_root() -> Path:
    return Path(os.environ.get("WINIT_DOWNLOAD_DIR", str(ROOT / "downloads")))


def _default_extract_parent() -> Path:
    return Path(os.environ.get("WINIT_EXTRACT_BASE", str(_downloads_root() / "extracted")))


def find_latest_export_zip(search_root: Path) -> Path:
    """优先 inventorySellerPortalExport*.zip，否则任意 .zip，按修改时间最新。"""
    pat = "inventorySellerPortalExport*.zip"
    cands = list(search_root.rglob(pat))
    if not cands:
        cands = list(search_root.rglob("*.zip"))
    if not cands:
        raise FileNotFoundError(
            f"在 {search_root} 下未找到 zip（可先运行 step02 下载，或检查 WINIT_DOWNLOAD_DIR）"
        )
    return max(cands, key=lambda p: p.stat().st_mtime)


def extract_zip(zip_path: Path, out_dir: Path) -> List[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    written: List[Path] = []
    with zipfile.ZipFile(zip_path, "r") as zf:
        for member in zf.infolist():
            if member.is_dir():
                continue
            zf.extract(member, path=out_dir)
            # extract 使用 member.filename，可能含子目录
            written.append(out_dir / member.filename)
    # 扁平列出实际生成的文件
    files = [p for p in out_dir.rglob("*") if p.is_file()]
    return sorted(files)


def _truncate_cell(v: object, max_len: int) -> str:
    if v is None:
        return ""
    s = str(v).replace("\n", " ").replace("\r", "")
    if len(s) > max_len:
        return s[: max_len - 1] + "…"
    return s


def preview_xlsx(
    xlsx_path: Path,
    *,
    head_rows: int,
    max_col_width: int,
    list_only: bool,
) -> None:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        print(f"文件：{xlsx_path.name}")
        print("工作表：", ", ".join(wb.sheetnames))
        if list_only:
            return
        for name in wb.sheetnames:
            ws = wb[name]
            print(f"\n--- 「{name}」前 {head_rows} 行（列宽至多 {max_col_width} 字符）---")
            for i, row in enumerate(ws.iter_rows(max_row=head_rows, values_only=True)):
                line = " | ".join(_truncate_cell(c, max_col_width) for c in row)
                print(line)
    finally:
        wb.close()


def export_sheets_to_csv(xlsx_path: Path, csv_dir: Path) -> List[Path]:
    """将各 sheet 导出为 UTF-8-BOM CSV，便于 Excel 打开。"""
    csv_dir.mkdir(parents=True, exist_ok=True)
    out_paths: List[Path] = []
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    try:
        for ws in wb.worksheets:
            safe = re.sub(r'[<>:"/\\|?*]', "_", ws.title) or "sheet"
            out = csv_dir / f"{safe}.csv"
            with out.open("w", encoding="utf-8-sig", newline="") as f:
                w = csv.writer(f)
                for row in ws.iter_rows(values_only=True):
                    w.writerow(["" if c is None else c for c in row])
            out_paths.append(out)
    finally:
        wb.close()
    return out_paths


def pick_xlsx_files(paths: Iterable[Path]) -> List[Path]:
    return sorted({p.resolve() for p in paths if p.suffix.lower() == ".xlsx"})


def main(argv: Optional[List[str]] = None) -> int:
    p = argparse.ArgumentParser(description="解压万邑通导出 zip 并预览 / 导出 xlsx")
    p.add_argument(
        "zip_path",
        nargs="?",
        default=None,
        help="zip 路径；省略则在 downloads 下自动选最新的 inventorySellerPortalExport*.zip",
    )
    p.add_argument(
        "--search-root",
        type=Path,
        default=None,
        help="自动查找 zip 时的根目录（默认 WINIT_DOWNLOAD_DIR 或 downloads/）",
    )
    p.add_argument(
        "--out",
        type=Path,
        default=None,
        help="解压目录（默认 <WINIT_EXTRACT_BASE>/<zip 主文件名>/）",
    )
    p.add_argument("--head", type=int, default=15, help="每个 sheet 预览行数（含表头）")
    p.add_argument(
        "--col-width",
        type=int,
        default=28,
        help="预览时每列最多字符数",
    )
    p.add_argument(
        "--list-only",
        action="store_true",
        help="只列出工作表名称，不打印行内容",
    )
    p.add_argument(
        "--export-csv",
        action="store_true",
        help="将 xlsx 各 sheet 导出为 CSV 到解压目录下的 csv/",
    )
    args = p.parse_args(argv)

    search_root = args.search_root or _downloads_root()
    if args.zip_path:
        zip_path = Path(args.zip_path).expanduser().resolve()
        if not zip_path.is_file():
            print("找不到 zip：", zip_path, file=sys.stderr)
            return 1
    else:
        try:
            zip_path = find_latest_export_zip(search_root)
        except FileNotFoundError as e:
            print(e, file=sys.stderr)
            return 1
        print("选用最新 zip：", zip_path, flush=True)

    out_dir = args.out
    if out_dir is None:
        out_dir = _default_extract_parent() / zip_path.stem
    else:
        out_dir = Path(out_dir).expanduser().resolve()

    extracted = extract_zip(zip_path, out_dir)
    xlsx_files = pick_xlsx_files(extracted)
    if not xlsx_files:
        print("解压完成，但未发现 .xlsx：", out_dir, file=sys.stderr)
        return 2

    print("解压到：", out_dir, flush=True)
    for xlsx in xlsx_files:
        preview_xlsx(
            xlsx,
            head_rows=max(1, args.head),
            max_col_width=max(8, args.col_width),
            list_only=args.list_only,
        )
        if args.export_csv:
            csv_dir = out_dir / "csv"
            outs = export_sheets_to_csv(xlsx, csv_dir)
            print(f"已导出 CSV（{len(outs)} 个）：", csv_dir, flush=True)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
