"""
从万邑通导出 xlsx（库存 Sheet）解析行并写入 inventory_daily。
首行表头为中文列名；整行保留为 row_json，并抽取常用字段做列索引。
"""

from __future__ import annotations

import json
import sqlite3
from pathlib import Path
from typing import Any, List, Tuple

import openpyxl

from winit_inventory_db import replace_snapshot_rows


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _cell_float(v: Any) -> Any:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def ingest_inventory_xlsx(
    conn: sqlite3.Connection,
    xlsx_path: Path,
    *,
    snapshot_date: str,
    account_id: int,
    account_username: str,
) -> int:
    """
    读取第一个工作表；第一行为表头。返回写入行数（先删后插，见 replace_snapshot_rows）。
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        header_row = next(it, None)
        if not header_row:
            return 0
        headers = [_cell_str(h) or f"col_{i}" for i, h in enumerate(header_row)]

        out_rows: List[Tuple[Any, ...]] = []
        for row in it:
            if row is None or all(c is None or c == "" for c in row):
                continue
            d: dict = {}
            for h, c in zip(headers, row):
                d[h] = c
            sku = _cell_str(d.get("商品编号"))
            if not sku:
                continue
            country = _cell_str(d.get("国家")) or None
            warehouse = _cell_str(d.get("仓库")) or None
            name_zh = _cell_str(d.get("中文名称")) or None
            name_en = _cell_str(d.get("英文名称")) or None
            qty_av = _cell_float(d.get("可用库存"))
            qty_oh = _cell_float(d.get("在库库存"))
            row_json = json.dumps(d, ensure_ascii=False, default=str)
            out_rows.append(
                (country, warehouse, sku, name_zh, name_en, qty_av, qty_oh, row_json)
            )
    finally:
        wb.close()

    return replace_snapshot_rows(
        conn,
        snapshot_date=snapshot_date,
        account_id=account_id,
        account_username=account_username,
        rows=out_rows,
    )
